from datetime import date


from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.chrome.options import Options


opt = Options()
opt.add_argument("--disable-infobars")
opt.add_argument("start-maximized")
opt.add_argument("--disable-extensions")
# Pass the argument 1 to allow and 2 to block
opt.add_experimental_option("prefs", {
    "profile.default_content_setting_values.media_stream_mic": 1,
    "profile.default_content_setting_values.media_stream_camera": 1,
    "profile.default_content_setting_values.geolocation": 1,
    "profile.default_content_setting_values.notifications": 1
  })


# ENTER THE PATH OF CHROME DRIVER HERE
PATH = "C:\chromedriver\chromedriver.exe"

driver = webdriver.Chrome(options=opt,executable_path=PATH)
driver.get('https://iiitlucknow.webex.com/webappng/sites/iiitlucknow/meeting/download/469b9f772066ddaf94f783b042528c9d')
driver.maximize_window()
browser = driver.find_element_by_link_text('Join from your browser.')
browser.click()
try:
    time.sleep(3)
    driver.switch_to.frame("thinIframe")
    driver.implicitly_wait(5)

    element = driver.find_element_by_xpath('//*[@id="meetingSimpleContainer"]/div[3]/div[2]/div/input')
    element.send_keys('LCS2019065_ATTENDANCE')
    element = driver.find_element_by_xpath('//*[@id="meetingSimpleContainer"]/div[3]/div[3]/div/input')
    element.send_keys('lcs2019065@iiitl.ac.in')
    element.send_keys(Keys.TAB)
    element.send_keys(Keys.RETURN)
    driver.implicitly_wait(5)
    time.sleep(10)

    mute = driver.find_element_by_xpath('//*[@id="meetingSimpleContainer"]/div[3]/div[2]/div[1]/div/button/span[2]')
    mute.click()
    video = driver.find_element_by_xpath('//*[@id="meetingSimpleContainer"]/div[3]/div[2]/div[2]/div/button/span[2]')
    video.click()
    join = driver.find_element_by_xpath('//*[@id="interstitial_join_btn"]')

    join.click()
    driver.implicitly_wait(10)
    time.sleep(35)

    party = driver.find_element_by_xpath('//*[@id="layoutdomid"]/div/div[3]/div[2]/div[3]/div/button[1]')
    party.click()


    element = driver.find_element_by_xpath(
        '//*[@id="layoutdomid"]/div/div[2]/div/div[3]/div/div[2]/div[2]/div/section/div[1]/div/div[1]')
    x= element.get_attribute('outerHTML')
except:
    print("not done")


dates = date.today()


import openpyxl as xl
wb = xl.load_workbook("attend.xlsx")
css= wb['Sheet1']
its= wb['Sheet2']



flag=0
roll_cs=["A"]*79
roll_it=["A"]*79
absentcs=[]
absentit=[]
presentit=0
presentcs=0
roll_it[0]=roll_cs[0]=str(dates)

x=x.lower()

for each in x:
    if each=='l':
        flag=1
        continue
    if flag==1:
        if each=="c":
            flag = 2
            continue
        elif each=="i":
            flag=-2
            continue
        else:
            flag = 0
            continue
    if flag == 2:
        if each=='s':
            flag=3
            continue
        else:
            flag=0
            continue
    if flag == -2:
        if each == "t":
            flag=-3
            continue
        else:
            flag=0
            continue

        # checking year
    if flag == -6:
        if each == '9':
            flag -= 1
            continue
        else:
            flag = 0
            continue
    if flag == 6:
        if each == '9':
            flag += 1
            continue
        else:
            flag = 0
            continue


    if flag==-8 and int(each)<10:
        temp=int(each)
        flag-=1
        continue
    if flag==-9:
        roll_it[temp * 10 + int(each)] = "P"
        flag=0
    if flag<=-3:
        flag-=1

    if flag==8:
        temp=int(each)
        flag+=1
        continue
    if flag==9:
        roll_cs[temp*10+int(each)]="P"
        flag=0
    if flag>=3:
        flag+=1

cn = css.max_column+1
cn2 = its.max_column+1
for i in range(0,78):
    if roll_cs[i+1]=="P":
        presentcs+=1
    if roll_it[i+1]=="P":
        presentit+=1
roll_cs[24]=roll_cs[31]=roll_cs[66]=""
for num in range(2, 81):
    if roll_cs[num-2]=="A":
        if num-2!=24 or num-2!=31 or num-2!=66:
            absentcs.append("LCS20190"+str(num-2))
    if roll_it[num - 2] == "A":
        absentit.append("LI20190"+str(num-2))
    css.cell(num, cn).value = roll_cs[num-2]
    its.cell(num, cn2).value = roll_it[num - 2]
wb.save('attend.xlsx')

print(f'''{dates} Attendance data

Total Students Present Today = {presentcs+presentit}

CS Students = {presentcs}
IT Students = {presentit}

CS Absenties
{absentcs}

IT Absenties 
{absentit}''')